import tkinter as tk
from tkinter import ttk
import pyttsx3
import speech_recognition as sr
import wikipedia
import pywhatkit
import webbrowser
import os
import datetime

import time
import re
from docx import Document
import math
import subprocess

from PIL import Image, ImageTk, ImageSequence, ImageDraw

# app_principal.py

import subprocess
import threading

# Variable global para almacenar el proceso de detección de ojos
proceso_deteccion = None

def iniciar_deteccion_ojos():
    global proceso_deteccion
    # Iniciar el proceso de detección de ojos en un hilo separado
    proceso_deteccion = subprocess.Popen(['python', 'deteccion_ojos.py'])

def detener_deteccion_ojos():
    global proceso_deteccion
    if proceso_deteccion:
        proceso_deteccion.terminate()  # Terminar el proceso de detección
        proceso_deteccion = None


# Configuración inicial del asistente
wake_word = "Soee"
listener = sr.Recognizer()
engine = pyttsx3.init()

voices = engine.getProperty('voices')
engine.setProperty('voice', voices[0].id)

# Función para que el asistente hable
def talk(text):
    print(text)
    engine.say(text)
    engine.runAndWait()

# Función para escuchar al usuario
def listen():
    with sr.Microphone() as source:
        print("Escuchando...")
        listener.adjust_for_ambient_noise(source, duration=0.7)
        audio = listener.listen(source, timeout=5)

    try:
        rec = listener.recognize_google(audio, language="es").lower()
        print("Texto detectado:", rec)
        return rec
    except sr.UnknownValueError:
        talk("No te entendí, intenta de nuevo")
        return None
    except sr.RequestError:
        talk("Error al conectar con el servicio de reconocimiento. Por favor, intenta de nuevo.")
        return None

# Función para la introducción inicial del asistente
def introduce():
    talk(f"Hola soy Soee, en que  puedo  ayudarte  hoy")


def despedida():
    try:
        mensaje = "Gracias por su atención. Muy pronto nos veremos con la tercera versión."
        talk(mensaje)
        # Espera a que termine de hablar antes de apagar
        os.system("shutdown /s /t 1")
    except Exception as e:
        print(f"Error al realizar la despedida: {str(e)}")
        talk("Ocurrió un error al intentar despedirme.")
# Función para reproducir música desde YouTube
def reproduce(rec):
    try:
        music = rec.replace('reproduce', '').strip()
        talk(f"Reproduciendo {music}")
        pywhatkit.playonyt(music)
    except Exception as e:
        print(f"Error al reproducir música: {str(e)}")
        talk("No pude reproducir la música.")

# Función para buscar en Wikipedia
from database import registerUser, getUser, register_search  # Asegúrate de importar la función register_search

# Modifica la función `busca` para incluir el registro en la base de datos
def busca(rec):
    try:
        search = rec.replace('busca', '').strip()
        wikipedia.set_lang("es")
        wiki = wikipedia.summary(search, sentences=1)
        talk(wiki)
        write_text(search + ": " + wiki)
        
        # Aquí deberías obtener el id del usuario actual. Por ejemplo, puedes almacenar el ID en una variable global si es necesario.
        id_usuario = 1  # Cambia esto por el método adecuado para obtener el id del usuario actual
        fecha_actual = datetime.datetime.now()
        register_search(id_usuario, search, fecha_actual)
        
    except wikipedia.exceptions.DisambiguationError as e:
        try:
            wiki = wikipedia.summary(e.options[0], sentences=1)
            talk(wiki)
            write_text(e.options[0] + ": " + wiki)
            
            # Aquí deberías obtener el id del usuario actual.
            id_usuario = 1
            fecha_actual = datetime.datetime.now()
            register_search(id_usuario, e.options[0], fecha_actual)
            
        except Exception as ex:
            print(f"Error al buscar en Wikipedia: {str(ex)}")
            talk("No pude encontrar información en Wikipedia.")
    except Exception as ex:
        print(f"Error al buscar en Wikipedia: {str(ex)}")
        talk("No pude encontrar información en Wikipedia.")

# Función para buscar en Google
def buscame(rec):
    try:
        something = rec.replace("búscame", '').strip()
        talk("Buscando " + something)
        webbrowser.open_new_tab(f"https://www.google.com/search?q={something}")
        write_text("Búsqueda en Google: " + something)
    except Exception as e:
        print(f"Error al buscar en el navegador: {str(e)}")
        talk("No pude buscar en el navegador.")

# Función para escribir texto en un archivo
def write_text(text):
    try:
        with open('output.txt', 'a', encoding='utf-8') as file:
            file.write(text + '\n')
        print("Texto escrito en el archivo 'output.txt'")
    except Exception as e:
        print(f"Error al escribir en el archivo: {str(e)}")

# Función para abrir aplicaciones
def abrir_aplicacion(rec):
    try:
        app = rec.replace('abre', '').strip()
        if "navegador" in app:
            os.system("start chrome")  # Cambia "chrome" por el navegador que uses
            talk(f"Abriendo {app}")
        elif "editor de texto" in app:
            os.system("start notepad")  # Puedes cambiar "notepad" por tu editor preferido
            talk(f"Abriendo {app}")
        elif "reproductor de música" in app:
            os.system("start wmplayer")  # Puedes cambiar "wmplayer" por tu reproductor preferido
            talk(f"Abriendo {app}")
        else:
            talk(f"No puedo abrir {app}.")
    except Exception as e:
        print(f"Error al abrir la aplicación: {str(e)}")
        talk(f"No pude abrir la aplicación {app}.")

# Función para mostrar el clima (sin implementación específica)
def muestra_clima(rec):
    try:
        location = rec.replace('muestra clima de', '').strip()
        talk(f"Mostrando el clima de {location}")
    except Exception as e:
        print(f"Error al consultar el clima: {str(e)}")
        talk("No pude consultar el clima en este momento.")

# Variable global para almacenar el proceso de detección de ojos
proceso_deteccion = None

# Función para iniciar la detección de ojos
def iniciar_deteccion_ojos():
    global proceso_deteccion
    if proceso_deteccion is None:
        proceso_deteccion = subprocess.Popen(['python', 'deteccion_ojos.py'])
        talk("Detección de ojos iniciada.")
    else:
        talk("La detección de ojos ya está en ejecución.")

# Función para detener la detección de ojos
def detener_deteccion_ojos():
    global proceso_deteccion
    if proceso_deteccion is not None:
        proceso_deteccion.terminate()
        proceso_deteccion = None
        talk("Detección de ojos detenida.")
    else:
        talk("No hay detección de ojos en ejecución.")


# Función para realizar cálculos matemáticos
def calcula(rec):
    try:
        rec = rec.lower()

        if 'promedio de' in rec:
            numbers = re.findall(r'\d+', rec)
            numbers = list(map(int, numbers))
            if numbers:
                resultado = sum(numbers) / len(numbers)
                talk(f"El promedio de {numbers} es {resultado}")
            else:
                talk("No pude encontrar números para calcular el promedio.")

        elif 'porcentaje de' in rec:
            match = re.search(r'(\d+)% de (\d+)', rec)
            if match:
                porcentaje = int(match.group(1))
                total = int(match.group(2))
                resultado = (porcentaje / 100) * total
                talk(f"El {porcentaje}% de {total} es {resultado}")
            else:
                talk("No pude encontrar el porcentaje y el total en el comando.")

        elif 'raíz cuadrada de' in rec:
            match = re.search(r'raíz cuadrada de (\d+)', rec)
            if match:
                numero = int(match.group(1))
                resultado = math.sqrt(numero)
                talk(f"La raíz cuadrada de {numero} es {resultado}")
            else:
                talk("No pude encontrar el número para calcular la raíz cuadrada.")

        elif 'potencia de' in rec:
            match = re.search(r'(\d+) elevado a (\d+)', rec)
            if match:
                base = int(match.group(1))
                exponente = int(match.group(2))
                resultado = base ** exponente
                talk(f"{base} elevado a {exponente} es {resultado}")
            else:
                talk("No pude encontrar la base y el exponente para calcular la potencia.")

        elif 'suma de' in rec:
            numbers = re.findall(r'\d+', rec)
            numbers = list(map(int, numbers))
            if numbers:
                resultado = sum(numbers)
                talk(f"La suma de {numbers} es {resultado}")
            else:
                talk("No pude encontrar números para sumar.")

        elif 'resta de' in rec:
            numbers = re.findall(r'\d+', rec)
            if len(numbers) == 2:
                minuendo = int(numbers[0])
                sustraendo = int(numbers[1])
                resultado = minuendo - sustraendo
                talk(f"La resta de {minuendo} menos {sustraendo} es {resultado}")
            else:
                talk("No pude encontrar dos números para restar.")

        elif 'multiplicación de' in rec:
            numbers = re.findall(r'\d+', rec)
            if len(numbers) == 2:
                factor1 = int(numbers[0])
                factor2 = int(numbers[1])
                resultado = factor1 * factor2
                talk(f"La multiplicación de {factor1} por {factor2} es {resultado}")
            else:
                talk("No pude encontrar dos números para multiplicar.")

        elif 'división de' in rec:
            numbers = re.findall(r'\d+', rec)
            if len(numbers) == 2:
                dividendo = int(numbers[0])
                divisor = int(numbers[1])
                if divisor != 0:
                    resultado = dividendo / divisor
                    talk(f"La división de {dividendo} entre {divisor} es {resultado}")
                else:
                    talk("No se puede dividir entre cero.")
            else:
                talk("No pude encontrar dos números para dividir.")

        else:
            talk("No entendí el tipo de cálculo que quieres realizar.")

    except Exception as e:
        print(f"Error al realizar el cálculo: {str(e)}")
        talk("No pude realizar el cálculo.")

# Función para configurar una alarma (sin implementación específica)
import datetime
import time
import re
import pygame  # Biblioteca para reproducir sonidos

# Inicializar pygame mixer
pygame.mixer.init()

# Cargar el archivo de sonido para la alarma
import pygame

def cargar_sonido():
    try:
        pygame.mixer.init()  # Inicializar el mezclador de pygame
        pygame.mixer.music.load('mi_alarma.mp3')  # Cambia el nombre del archivo según corresponda
    except pygame.error as e:
        print(f"No se pudo cargar el archivo de sonido: {e}")

def reproducir_sonido():
    pygame.mixer.music.play()  # Reproducir el sonido


# Reproducir el sonido

def configurar_alarma(rec):
    try:
        # Buscar la hora en el comando con el nuevo formato
        match = re.search(r'configura alarma para las (\d+):(\d+)', rec)
        if match:
            hora_alarma = int(match.group(1))
            minuto_alarma = int(match.group(2))

            # Obtener la hora y el minuto actuales
            ahora = datetime.datetime.now()
            alarma_tiempo = ahora.replace(hour=hora_alarma, minute=minuto_alarma, second=0, microsecond=0)

            # Si la alarma está programada para un tiempo pasado, configurar para el próximo día
            if alarma_tiempo < ahora:
                alarma_tiempo += datetime.timedelta(days=1)

            # Calcular el tiempo restante hasta la alarma
            tiempo_espera = (alarma_tiempo - ahora).total_seconds()

            # Informar al usuario y esperar
            talk(f"Alarma configurada para las {hora_alarma:02d}:{minuto_alarma:02d}.")
            time.sleep(tiempo_espera)  # Esperar hasta el momento de la alarma

            # Reproducir el sonido de la alarma
            
            talk("¡Es hora!")
            print('¡Suena alarma!')

        else:
            talk("No pude entender la hora para la alarma. Usa el formato 'configura alarma para las HH:MM'.")

    except Exception as e:
        print(f'Error al configurar alarma: {str(e)}')
        talk('No pude configurar la alarma.')

# Cargar el sonido de la alarma al inicio
cargar_sonido()


# Función para decir la hora actual
def que_hora_es():
    """Función para decir la hora actual."""
    try:
        hora_actual = datetime.datetime.now().strftime("%H:%M")
        talk(f"Son las {hora_actual}")
    except Exception as e:
        print(f"Error al obtener la hora actual: {str(e)}")
        talk("No pude obtener la hora actual.")

def que_fecha_es():
    """Función para decir la fecha actual."""
    try:
        fecha_actual = datetime.datetime.now().strftime("%d de %B de %Y")
        talk(f"Hoy es {fecha_actual}")
    except Exception as e:
        print(f"Error al obtener la fecha actual: {str(e)}")
        talk("No pude obtener la fecha actual.")
# Función para transcribir dictado a texto
from docx import Document
import speech_recognition as sr

listener = sr.Recognizer()

def transcribir_texto():
    try:
        talk("Por favor, dicta el texto que deseas transcribir.")
        with sr.Microphone() as source:
            print("Escuchando para dictado...")
            listener.adjust_for_ambient_noise(source, duration=0.7)
            audio = listener.listen(source, timeout=7)  # Ajusta el timeout si es necesario

        try:
            text = listener.recognize_google(audio, language="es")
            talk(f"He transcrito el siguiente texto: {text}")
            
            # Crear un nuevo documento de Word
            doc = Document()
            doc.add_paragraph(text)
            
            doc.save('Word.docx')
            
            print("Texto transcrito y guardado en 'nota.docx'")
        except sr.UnknownValueError:
            talk("No pude entender el dictado, intenta de nuevo.")
        except sr.RequestError as e:
            talk("No pude conectar con el servicio de reconocimiento.")
    except Exception as e:
        print(f"Error al transcribir el texto: {str(e)}")
        talk("Ocurrió un error al transcribir el texto.")
        

# Función principal para iniciar el asistente y la interfaz gráfica
def iniciar_bot():
    introduce()
    while True:
        try:
            rec = listen()
            if rec is not None:
                if 'reproduce' in rec:
                    reproduce(rec)
                elif 'busca' in rec:
                    busca(rec)
                elif 'búscame' in rec:
                    buscame(rec)
                elif 'abre' in rec:
                    abrir_aplicacion(rec)
                elif 'muestra clima de' in rec:
                    muestra_clima(rec)
                elif 'calcula' in rec:
                    calcula(rec)
                elif 'configura alarma' in rec:
                    configurar_alarma(rec)
                elif 'qué hora es' in rec:
                    que_hora_es()
                elif 'qué fecha es' in rec:
                    que_fecha_es()
                elif 'transcribe' in rec:
                    transcribir_texto()
                elif 'detente' in rec:
                    talk("Bot detenido.")
                    break
                elif 'inicia detección de ojos' in rec:
                    iniciar_deteccion_ojos()
                elif 'detén detección de ojos' in rec:
                    detener_deteccion_ojos()
                    talk("Lo siento, no entendí ese comando.")
                
                elif 'despedida' in rec:
                    despedida()   
        except Exception as e:
            talk("Ocurrió un error, intenta de nuevo.") 
            print(f"Ocurrió un error: {str(e)}")

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import openpyxl
from datetime import datetime
import os
import speech_recognition as sr
import pyttsx3
import webbrowser

# Inicialización del motor de texto a voz
engine = pyttsx3.init()

# Función para guardar registros en archivos Excel
def guardar_en_excel(archivo, datos):
    try:
        wb = openpyxl.load_workbook(archivo)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Elemento', 'Fecha', 'Hora'])
    else:
        ws = wb.active

    fecha_hora = datetime.now()
    fecha = fecha_hora.strftime('%Y-%m-%d')
    hora = fecha_hora.strftime('%H:%M:%S')
    ws.append([datos, fecha, hora])
    wb.save(archivo)

# Función para agregar archivos
def agregar_archivo():
    archivo = filedialog.askopenfilename(title="Selecciona un archivo")
    if archivo:
        guardar_en_excel('ver_archivos.xlsx', archivo)
        messagebox.showinfo("Archivo agregado", f"Archivo guardado:\n{archivo}")

# Función para agregar aplicaciones
def agregar_aplicacion():
    aplicacion = filedialog.askopenfilename(title="Selecciona una aplicación")
    if aplicacion:
        guardar_en_excel('ver_aplicaciones.xlsx', aplicacion)
        messagebox.showinfo("Aplicación agregada", f"Aplicación guardada:\n{aplicacion}")

# Función para agregar páginas web
def agregar_pagina():
    pagina = simpledialog.askstring("Agregar Página Web", "Ingresa la URL de la página web:")
    if pagina:
        guardar_en_excel('ver_paginas.xlsx', pagina)
        messagebox.showinfo("Página Web agregada", f"Página Web guardada:\n{pagina}")

# Función para mostrar archivos guardados
def ver_archivos():
    archivo = 'ver_archivos.xlsx'
    if os.path.exists(archivo):
        os.startfile(archivo)
    else:
        messagebox.showinfo("Archivos guardados", "No hay archivos guardados.")

# Función para mostrar aplicaciones guardadas
def ver_aplicaciones():
    archivo = 'ver_aplicaciones.xlsx'
    if os.path.exists(archivo):
        os.startfile(archivo)
    else:
        messagebox.showinfo("Aplicaciones guardadas", "No hay aplicaciones guardadas.")

# Función para mostrar páginas web guardadas
def ver_paginas():
    archivo = 'ver_paginas.xlsx'
    if os.path.exists(archivo):
        os.startfile(archivo)
    else:
        messagebox.showinfo("Páginas Web guardadas", "No hay páginas guardadas.")

# Función para buscar y abrir un archivo específico por nombre
def abrir_archivo_por_nombre(nombre_archivo):
    archivo_excel = 'ver_archivos.xlsx'
    if not os.path.exists(archivo_excel):
        engine.say("No hay archivos guardados.")
        engine.runAndWait()
        return
    
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        archivo_guardado = row[0]  # Suponiendo que el archivo está en la primera columna
        if archivo_guardado and nombre_archivo.lower() in archivo_guardado.lower():
            if os.path.exists(archivo_guardado):
                os.startfile(archivo_guardado)
                engine.say(f"Abriendo {archivo_guardado}")
                engine.runAndWait()
                return
            else:
                engine.say(f"El archivo {archivo_guardado} ya no existe.")
                engine.runAndWait()
                return
    
    engine.say(f"No se encontró un archivo con el nombre {nombre_archivo}")
    engine.runAndWait()

# Función para abrir el primer archivo guardado
def abrir_primer_archivo():
    archivo_excel = 'ver_archivos.xlsx'
    if not os.path.exists(archivo_excel):
        engine.say("No hay archivos guardados.")
        engine.runAndWait()
        return
    
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        archivo_guardado = row[0]  # Suponiendo que el archivo está en la primera columna
        if archivo_guardado and os.path.exists(archivo_guardado):
            os.startfile(archivo_guardado)
            engine.say(f"Abriendo {archivo_guardado}")
            engine.runAndWait()
            return
    
    engine.say("No hay archivos disponibles para abrir.")
    engine.runAndWait()

# Función para reconocer comandos de voz
def reconocer_comando():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        engine.say("Escuchando...")
        engine.runAndWait()
        audio = r.listen(source)
        try:
            comando = r.recognize_google(audio, language="es-ES").lower()  # Configuración del idioma a español
            engine.say(f"Comando recibido: {comando}")
            engine.runAndWait()
            procesar_comando(comando)
        except sr.UnknownValueError:
            engine.say("No entendí el comando, por favor repite.")
            engine.runAndWait()
        except sr.RequestError:
            engine.say("No se pudo conectar con el servicio de reconocimiento de voz.")
            engine.runAndWait()

# Función para procesar el comando de voz
def procesar_comando(comando):
    if 'abre archivo' in comando:
        nombre_archivo = comando.replace('abre archivo', '').strip()
        if nombre_archivo:
            abrir_archivo_por_nombre(nombre_archivo)
        else:
            abrir_primer_archivo()  # Abre el primer archivo disponible si no se proporciona un nombre específico
    elif 'abre aplicación' in comando:
        ver_aplicaciones()
    elif 'abre página' in comando:
        nombre_pagina = comando.replace('abre página', '').strip()
        if nombre_pagina:
            abrir_pagina_por_nombre(nombre_pagina)
        else:
            ver_paginas()
    else:
        engine.say("Comando no reconocido.")
        engine.runAndWait()

# Función para abrir una página web específica por nombre
def abrir_pagina_por_nombre(nombre_pagina):
    archivo_excel = 'ver_paginas.xlsx'
    if not os.path.exists(archivo_excel):
        engine.say("No hay páginas web guardadas.")
        engine.runAndWait()
        return
    
    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        pagina_guardada = row[0]  # Suponiendo que la URL está en la primera columna
        if pagina_guardada and nombre_pagina.lower() in pagina_guardada.lower():
            webbrowser.open(pagina_guardada)
            engine.say(f"Abriendo página web {pagina_guardada}")
            engine.runAndWait()
            return
    
    engine.say(f"No se encontró una página web con el nombre {nombre_pagina}")
    engine.runAndWait()            

# Función para crear una imagen redonda
def make_round_image(image_path, size):
    try:
        # Cargar la imagen original
        original_image = Image.open(image_path).convert("RGBA")
        
        # Redimensionar la imagen para ajustarla al tamaño deseado
        original_image = original_image.resize(size, resample=Image.LANCZOS)
        
        # Crear una máscara circular del tamaño de la imagen redimensionada
        mask = Image.new("L", size, 0)
        draw = ImageDraw.Draw(mask)
        draw.ellipse((0, 0) + size, fill=255)
        
        # Crear una imagen RGBA completamente transparente
        round_image = Image.new("RGBA", size, (0, 0, 0, 0))
        
        # Aplicar la máscara a la imagen original para recortarla en forma de círculo
        round_image.paste(original_image, (0, 0), mask=mask)
        
        return round_image
    
    except IOError:
        print(f"No se pudo cargar la imagen desde '{image_path}'. Verifica la ruta y el formato.")
        return None

def button_click():
    talk("")
    rec = listen()
    if rec:
        iniciar_bot()  # Iniciar la interacción del bot basada en el comando recibido

def crear_imagen_degradado(color1, color2, width, height):
    """Crea una imagen PNG con un degradado vertical."""
    gradient = Image.new('RGB', (width, height))
    for y in range(height):
        r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
        r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
        r = int(r1 + (r2 - r1) * y / height)
        g = int(g1 + (g2 - g1) * y / height)
        b = int(b1 + (b2 - b1) * y / height)
        gradient.putpixel((0, y), (r, g, b))
    return gradient

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("IA SOEE")

# Obtener el tamaño de la pantalla
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calcular las dimensiones y posición de la ventana
window_width = 1080
window_height = 768
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# Establecer la geometría de la ventana
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Establecer el color de fondo de la ventana
root.configure(bg='black')

# Cargar la imagen estática
image_path = "portada_Soee.jpg"  # Ruta a la imagen estática
try:
    image = Image.open(image_path)
    image = image.resize((window_width, window_height))  # Ajustar el tamaño de la imagen si es necesario
    image_tk = ImageTk.PhotoImage(image)
except IOError:
    print(f"No se pudo cargar la imagen desde '{image_path}'. Verifica la ruta y el formato.")
    image_tk = None

# Mostrar la imagen estática en un Label
if image_tk:
    image_label = tk.Label(root, image=image_tk, bg='black')
    image_label.place(x=0, y=0, relwidth=1, relheight=1)  # Ajustar posición y tamaño de la imagen

# Crear una imagen de degradado para los botones
degradado_img = crear_imagen_degradado('#19C6C8', '#0d9b9e', 200, 50)
degradado_img_tk = ImageTk.PhotoImage(degradado_img)

# Función para crear botones con fondo de degradado
def crear_boton(canvas, text, command):
    """Crea un botón con fondo de degradado."""
    boton = tk.Button(canvas, text=text, command=command, image=degradado_img_tk, compound='center', relief='flat', bd=0, highlightthickness=0, font=('Helvetica', 12, 'bold'), fg='white', bg='#19C6C8')
    boton.pack(pady=10, fill=tk.X, padx=0)  # Ajustar el relleno y el margen
    return boton

# Crear un marco para los botones a la izquierda
frame_botones = tk.Frame(root, bg='black')
frame_botones.pack(side=tk.LEFT, fill=tk.Y, padx=0, pady=20)

# Crear un canvas para los botones
canvas = tk.Canvas(frame_botones, width=200, height=window_height, bg='black', bd=0, highlightthickness=0)
canvas.pack(fill=tk.Y, side=tk.LEFT)

# Crear los botones con el fondo de degradado
btn_agregar_archivo = crear_boton(canvas, "Agregar Archivo", agregar_archivo)
btn_agregar_aplicacion = crear_boton(canvas, "Agregar Aplicación", agregar_aplicacion)
btn_agregar_pagina = crear_boton(canvas, "Agregar Página Web", agregar_pagina)
btn_ver_archivos = crear_boton(canvas, "Mostrar Archivos", ver_archivos)
btn_ver_aplicaciones = crear_boton(canvas, "Mostrar Aplicaciones", ver_aplicaciones)
btn_ver_paginas = crear_boton(canvas, "Mostrar Páginas", ver_paginas)
btn_reconocer_comando = crear_boton(canvas, "Reconocer Comando de Voz", reconocer_comando)

# Botón para interactuar con el asistente
button = tk.Button(root, text="Habla con Soee", bg='#19C6C8', fg='white', font=('Helvetica', 12, 'bold'), command=button_click, relief='flat', bd=0)
button.pack(pady=20)

# Ejecutar la interfaz gráfica
root.mainloop()
